import openpyxl
import pynlpir


class LocationRecognition:
    def __init__(self):
        # data path
        self.data_path = '../data/ECM_data_baidu.xlsx'

        # nlpir 初始化
        pynlpir.open()

        # 行政区划path
        self.adm_excel_path = '../data/全国行政区划分级数据_2018.xlsx'
        self.adm_sheet = self.load_excel(self.adm_excel_path)['t_position']
        self.adm_cityname = self.load_col_from_sheet('C', self.adm_sheet)
        self.adm_code = self.load_col_from_sheet('A', self.adm_sheet)
        self.adm_dep_code = self.load_col_from_sheet('B', self.adm_sheet)

    def extract_file_location(self):
        # step 1: 载入数据
        workbook = self.load_excel(self.data_path)
        sheet_list = self.load_sheet_list(workbook)

        # step 2:逐个表抽取地名
        for sheet in sheet_list:
            content_list = self.get_contentlist_from_sheet(sheet)
            row = 2
            for content in content_list:
                if content:
                    # get toponyms
                    toponyms, toponyms_set = self.extract_toponym_list_from_content(content)
                    sheet.cell(row=row, column=12, value=str(toponyms))
                    sheet.cell(row=row, column=13, value=str(toponyms_set))

                    frequency_dic = self.toponym_frequency(toponyms, toponyms_set)
                    sheet.cell(row=row, column=14, value=str(frequency_dic))

                    t1, t2, t3 = self.determine_toponyms_by_fre_dict(frequency_dic)
                    sheet.cell(row=row, column=15, value=str(t1))
                    sheet.cell(row=row, column=16, value=str(t2))
                    sheet.cell(row=row, column=17, value=str(t3))
                print(sheet.title + ' in :'+str(row) + ' has been processed.')
                row = row + 1
            print(sheet.title + ' has been processed.')
            workbook.save(self.data_path)

    def determine_toponyms_by_fre_dict(self, freq_dict):
        t1 = ''
        t2 = ''
        t3 = ''
        if freq_dict:
            if len(freq_dict) > 1:
                # dict 按照 地名次数排序
                counts = freq_dict.values()
                counts_list = list(counts)
                counts_list.sort()
                t1_value = counts_list[len(counts_list) - 1]
                t2_value = counts_list[len(counts_list) - 2]
                for ts in freq_dict:
                    if t1_value is freq_dict[ts]:
                        t1 = ts
                    elif t2_value is freq_dict[ts]:
                        t2 = ts
            else:
                for t in freq_dict:
                    t1 = t

        if t1 in freq_dict:
            del freq_dict[t1]
        if t2 in freq_dict:
            del freq_dict[t2]
        for f_tp in freq_dict:
            if self.adm_belong_to(f_tp, t1):
                t3 = f_tp
        if t3 == '':
            t3 = t1

        return t1, t2, t3

    def adm_belong_to(self, sub_pn, top_pn):
        belong = False
        if sub_pn in self.adm_cityname:
            index_row = 0
            for city_name in self.adm_cityname:
                if sub_pn in city_name:
                    dep_code = self.adm_dep_code[index_row]
                    dep_list = self.find_dep_list(dep_code)
                    if dep_list:
                        for up_tp in dep_list:
                            if top_pn in up_tp:
                                belong = True
                index_row = index_row + 1
        return belong

    def find_dep_list(self, code):
        dep_list = []
        dep_set = set()
        if code[-1] != 'x':
            code_list = list(code)
            code_list[4] = '0'
            code_list[5] = '0'
            code_list[6] = '0'
            code_list[7] = '0'
            code_list[8] = '0'
            dc1 = ''.join(code_list)
            code_list[2] = '0'
            code_list[3] = '0'
            dc2 = ''.join(code_list)
            code_list[8] = 'x'
            dc3 = ''.join(code_list)
            # 得到地名
            index_row = 0
            for dep_code in self.adm_dep_code:
                if dep_code == dc1:
                    dc1 = self.adm_cityname[index_row]
                if dep_code == dc2:
                    dc2 = self.adm_cityname[index_row]
                if dep_code == dc3:
                    dc3 = self.adm_cityname[index_row]
                index_row = index_row + 1
            dep_set.add(dc1)
            dep_set.add(dc2)
            dep_set.add(dc3)
        if dep_set:
            for t in dep_set:
                dep_list.append(t)
        return dep_list

    def toponym_frequency(self, toponyms_list, toponyms_set):
        tpm_fre_dic = {}
        for t in toponyms_set:
            tpm_fre_dic[t] = toponyms_list.count(t)
        return tpm_fre_dic

    def extract_toponym_list_from_content(self, content):
        toponyms = []
        toponym_set = set()
        # 逐句抽取地名
        seg_list = pynlpir.segment(content, pos_names='raw')
        for word_pair in seg_list:
            if word_pair[1] == "ns":
                placename = word_pair[0].replace("\n", "")
                if placename != '中国' and placename != '长江' and placename != '华北' and placename != '中原' and placename != '江南' and placename != '东北' and placename != '西南' and placename != '黄河':
                    toponyms.append(placename)
                    toponym_set.add(placename)
        # 检测缩略地名  e.g.枣庄/枣庄市
        filtered_toponym_list, filtered_toponym_set = self.replace_abb_toponyms(toponyms, toponym_set)

        return filtered_toponym_list, filtered_toponym_set

    def replace_abb_toponyms(self, toponyms_list, toponyms_set):
        abb = {}
        filtered_toponym_list = []
        filtered_toponym_set = set()
        place_name = ''
        full_place_name = ''
        # 检测缩略地名  e.g.枣庄/枣庄市
        for place_name in toponyms_set:
            for full_place_name in toponyms_set:
                if place_name in full_place_name and place_name != full_place_name:
                    abb[place_name] = full_place_name
        if abb:
            for t in toponyms_list:
                if t in abb.keys():
                    filtered_toponym_list.append(abb[t])
                else:
                    filtered_toponym_list.append(t)
        else:
            filtered_toponym_list = toponyms_list
            filtered_toponym_set = toponyms_set
        for ft in filtered_toponym_list:
            filtered_toponym_set.add(ft)
        return filtered_toponym_list, filtered_toponym_set

    def load_excel(self, file_path):
        try:
            wb = openpyxl.load_workbook(file_path)
            print("data file read successfully.")
            return wb
        except Exception as res:
            print("data file read unsuccessfully.")
            print(res)

    def load_sheet_list(self, wb):
        sheets = []
        if wb:
            shs = wb.get_sheet_names()
            for i in range(len(shs)):
                sheets.append(wb.get_sheet_by_name(shs[i]))
        return sheets

    def get_contentlist_from_sheet(self, sheet):
        content_list = []
        for i in range(sheet.max_row + 1):
            if i > 1:
                # read each row
                content_list.append(str(sheet.cell(row=i, column=7).value))
        return content_list

    def load_col_from_sheet(self, col_code, sheet):
        cols = []
        col = sheet[col_code]
        for x in range(1, len(col)):
            cols.append(str(col[x].value))
        return cols


lr = LocationRecognition()
lr.extract_file_location()
