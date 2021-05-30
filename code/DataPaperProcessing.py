from EcoCivMdl.code import Utility
import openpyxl
import re
from translate import Translator
import requests
from fake_useragent import UserAgent
from bs4 import BeautifulSoup
import random
import urllib
from lxml import etree
from hashlib import md5
import time
from EcoCivMdl.code import Baidu_trans


class DataPaperProcessing:

    def __init__(self):
        # 通用function
        self.ul = Utility.Utility_fc()

        # mode-code pair path
        self.data_path = '../data/CEA_Distribution_2018_2020.xlsx'

        # 载入数据
        self.wb = self.ul.load_excel(self.data_path)
        self.ws = self.wb["数据集"]

        # fake agent
        self.ua = UserAgent()
        self.baiduTrans = Baidu_trans.Baidu_trans()

    def process_logic(self):

        # # 找到空缺经纬度
        # missing_data = self.find_missing_data(ws)
        #
        # # 填充空缺
        # self.fill_data(ws, missing_data)
        # wb.save(self.data_path)

        # 翻译所需列
        self.translate_cols()

    def translate_cols(self):
        # translator setting
        translator = Translator(from_lang="chinese", to_lang="english")
        # print(translator.translate("笔"))

        traslate_col_list = ["D", "F", "K", "L"]
        traslate_col_out_list = ["R", "S", "T", "U", "V"]
        # traslate_col_list = ["D"]

        # 阻止睡眠，加速trans
        last_word = ''
        last_word_trams = ''
        trans_result = ''

        for t_col_id in range(len(traslate_col_list)):
            # read list
            tra_list = self.get_col_list_from_sheet(self.ws, traslate_col_list[t_col_id])

            # loop list
            for item_num in range(len(tra_list)):
                # translate process
                # trans_result = translator.translate(tra_list[item_num])                     # translate python 翻译
                # trans_result = self.translate_youdao(tra_list[item_num])                    # youdao在线
                print(tra_list[item_num])
                # trans_result = self.youdao.translate(tra_list[item_num])                     # 有道破解秘钥
                if tra_list[item_num] != last_word:
                    trans_result = self.baiduTrans.translate_baidu(tra_list[item_num])
                else:
                    trans_result = last_word_trams

                # set last word
                last_word = tra_list[item_num]
                last_word_trams = trans_result

                # write translated cell
                self.ws.cell(row=item_num + 2, column=18 + t_col_id, value=trans_result)
                print(traslate_col_list[t_col_id] + "——" + str(item_num + 1) + " of " + str(
                    len(tra_list)) + "——has finished...")

            # save each translate results
            self.wb.save(self.data_path)

    def translate_youdao(self, word):
        headers = {
            "User-Agent": self.ua.random,
            'Connection': 'close',
        }
        # 有道词典 api
        url = 'http://fanyi.youdao.com/translate?smartresult=dict&smartresult=rule&smartresult=ugc&sessionFrom=null'
        # 传输的参数，其中 i 为需要翻译的内容
        key = {
            'type': "AUTO",
            'i': word,
            "doctype": "json",
            "version": "2.1",
            "keyfrom": "fanyi.web",
            "ue": "UTF-8",
            "action": "FY_BY_CLICKBUTTON",
            "typoResult": "true"
        }

        # 随机IP
        ip_list = self.get_ip_list()
        proxies = self.get_random_ip(ip_list)
        print(proxies)

        # key 这个字典为发送给有道词典服务器的内容
        response = requests.post(url, proxies=proxies, headers=headers, data=key)
        # 判断服务器是否相应成功
        if response.status_code == 200:
            # 然后相应的结果
            print(response.text)
            return response.text
        else:
            print("有道词典调用失败")
            # 相应失败就返回空
            return None

    # 获取salt,sign,ts
    def get_salt_sign_ts(self, word):
        # ts
        ts = str(int(time.time() * 1000))
        # salt
        salt = ts + str(random.randint(0, 9))
        # sign
        string = "fanyideskweb" + word + salt + "n%A-rKaT5fb[Gy?;N5@Tj"
        s = md5()
        s.update(string.encode())
        sign = s.hexdigest()

        return salt, sign, ts

    # 主函数有道翻译
    def attack_yd(self, word):
        # 1. 先拿到salt,sign,ts
        salt, sign, ts = self.get_salt_sign_ts(word)
        # 2. 定义form表单数据为字典: data={}
        # 检查了salt sign
        data = {
            "i": word,
            "from": "AUTO",
            "to": "AUTO",
            "smartresult": "dict",
            "client": "fanyideskweb",
            "salt": salt,
            "sign": sign,
            "ts": ts,
            "bv": "7e3150ecbdf9de52dc355751b074cf60",
            "doctype": "json",
            "version": "2.1",
            "keyfrom": "fanyi.web",
            "action": "FY_BY_REALTlME",
        }
        # 3. 直接发请求:requests.post(url,data=data,headers=xxx)
        headers = {
            # 检查频率最高 - 3个
            "Cookie": "OUTFOX_SEARCH_USER_ID=970246104@10.169.0.83; OUTFOX_SEARCH_USER_ID_NCOO=570559528.1224236; _ntes_nnid=96bc13a2f5ce64962adfd6a278467214,1551873108952; JSESSIONID=aaae9i7plXPlKaJH_gkYw; td_cookie=18446744072941336803; SESSION_FROM_COOKIE=unknown; ___rl__test__cookies=1565689460872",
            "Referer": "http://fanyi.youdao.com/",
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/76.0.3809.100 Safari/537.36",
        }
        html = requests.post(
            url="'https://fanyi.youdao.com/translate_o?smartresult=dict&smartresult=rule",
            data=data,
            headers=headers
        ).json()
        # res.json() 将json格式的字符串转为python数据类型
        result = html['translateResult'][0][0]['tgt']

        print(result)
        return result

    # 将获取到的代理IP存到ip_list中并返回列表
    def get_ip_list(self):
        headers = {
            "User-Agent": self.ua.random,
            'Connection': 'close',
        }

        page = random.randint(1, 4000)

        request = urllib.request.Request("https://www.kuaidaili.com/free/inha/" + str(page), headers=headers)
        html = urllib.request.urlopen(request).read()

        ip_list = []

        # 解析网页信息，从中提取代理 ip 的数据
        content = etree.HTML(html)
        ip = content.xpath('//td[@data-title="IP"]/text()')
        port = content.xpath('//td[@data-title="PORT"]/text()')

        # 将代理 ip 信息存入 proxyList 列表
        for i in ip:
            for p in port:
                ip_list.append(i + ':' + p)

        return ip_list

    # 从代理IP列表中随机取出一个IP并返回
    def get_random_ip(self, ip_list):
        proxy_list = []
        for ip in ip_list:
            proxy_list.append('http://' + ip)
        proxy_ip = random.choice(proxy_list)
        proxies = {'http': proxy_ip}
        return proxies

    def fill_data(self, ws, missing_data):
        for i in range(len(missing_data)):
            missing_data_toponym = ws["J"][missing_data[i]].value

            # 找已有坐标
            x, y = self.find_topoynm_xy(ws, missing_data_toponym)
            # 写入坐标
            if x != '':
                ws.cell(row=missing_data[i] + 1, column=11, value=x)
                ws.cell(row=missing_data[i] + 1, column=12, value=y)

            print(str(i) + " in " + str(len(missing_data)) + " has been processed,")

    # 找地名对应的xy
    def find_topoynm_xy(self, ws, toponym):
        location_list = self.get_col_list_from_sheet(ws, "J")
        LNG_list = self.get_col_list_from_sheet(ws, "K")
        LAT_list = self.get_col_list_from_sheet(ws, "L")
        x = ''
        y = ''
        for i in range(len(location_list)):
            if toponym == location_list[i] and LNG_list[i] != '':
                x = LNG_list[i]
                y = LAT_list[i]
                break
        return x, y

    def find_missing_data(self, work_sheet):
        # 获取lng列
        LNG_list = self.get_col_list_from_sheet(work_sheet, "K")

        missing_data_id_list = []

        for i in range(len(LNG_list)):
            if LNG_list[i] == "":
                missing_data_id_list.append(i + 1)
        print("get missing data id list successfully.")
        return missing_data_id_list

    def get_col_list_from_sheet(self, sheet, col_code):
        col = sheet[col_code]
        datalist = []
        for x in range(len(col)):
            if x > 0:
                value = col[x].value
                if value is None:
                    datalist.append("")
                else:
                    value = str(value)
                    datalist.append(value)
        return datalist


dpp = DataPaperProcessing()
dpp.process_logic()
