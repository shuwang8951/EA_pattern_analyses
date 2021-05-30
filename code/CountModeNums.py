from EcoCivMdl.code import Utility
import openpyxl


class CountModeNums:

    def __init__(self):
        # 通用function
        self.ul = Utility.Utility_fc()

        # data path
        # self.data_path = '../data/ECM_clusterdata_xy-map-simple-value.xlsx'         #top10file
        # self.data_path = '../data/ECM_clusterdata_xy-map-3level.xlsx'                 #2、3级
        # self.data_path = '../data/Internet+.xlsx'                                     #分时间
        self.data_path = '../data/forest-crop-sparse.xlsx'

    def processes(self):
        # step 1: 载入top10数据
        workbook = self.ul.load_excel(self.data_path)
        sheets = self.ul.load_sheet_list(workbook)

        # step 2: 单个mode计算
        for sheet in sheets:
            print(sheet.title + " has been processing...")
            count_sheet = workbook.create_sheet(sheet.title + "_count")
            self.count_mode_num(workbook, sheet, count_sheet)
            print(sheet.title + " has finished.")

    def count_mode_num(self, workbook, sheet, count_sheet):

        sheet_max_row = sheet.max_row

        id_list = self.ul.get_col_list_from_sheet(sheet, "A")

        # 计算record记录数
        agg_num = 1

        de_id_list = []

        for index in range(2, sheet_max_row + 1):
            print(sheet.title + "    " + str(index) + " has been processing...")
            id = str(sheet.cell(row=index, column=1).value)
            ori_id=id
            title = str(sheet.cell(row=index, column=2).value)
            date = str(sheet.cell(row=index, column=3).value)

            short_s = str(sheet.cell(row=index, column=4).value)
            long_s = str(sheet.cell(row=index, column=5).value)
            mode = str(sheet.cell(row=index, column=6).value)

            site = str(sheet.cell(row=index, column=7).value)

            longitude = str(sheet.cell(row=index, column=8).value)
            latitude = str(sheet.cell(row=index, column=9).value)

            count = 1

            # 若不在删除列表中——重复列表
            if id not in de_id_list:
                flow_row = ""
                for flow_id in id_list:
                    # get flow row
                    for row in range(1, sheet_max_row + 1):
                        if flow_id == str(sheet.cell(row=row, column=1).value) and flow_id != ori_id:
                            flow_row = row

                            # 存在相同空间位置
                            if longitude == str(sheet.cell(row=flow_row, column=8).value) and latitude == str(
                                    sheet.cell(row=flow_row, column=9).value):
                                id = id + "@@" + str(sheet.cell(row=flow_row, column=1).value)
                                title = title + "@@" + str(sheet.cell(row=flow_row, column=2).value)
                                date = date + "@@" + str(sheet.cell(row=flow_row, column=3).value)
                                mode = mode + "@@" + str(sheet.cell(row=flow_row, column=6).value)
                                short_s = short_s + "@@" + str(sheet.cell(row=flow_row, column=4).value)
                                long_s = long_s + "@@" + str(sheet.cell(row=flow_row, column=5).value)
                                site = site + "@@" + str(sheet.cell(row=flow_row, column=7).value)
                                count = count + 1
                                de_id_list.append(flow_id)

                # 存储记录
                count_sheet.cell(row=agg_num, column=1, value=id)
                count_sheet.cell(row=agg_num, column=2, value=site)
                count_sheet.cell(row=agg_num, column=3, value=longitude)
                count_sheet.cell(row=agg_num, column=4, value=latitude)
                count_sheet.cell(row=agg_num, column=5, value=mode)
                count_sheet.cell(row=agg_num, column=6, value=title)
                count_sheet.cell(row=agg_num, column=7, value=date)
                count_sheet.cell(row=agg_num, column=8, value=short_s)
                count_sheet.cell(row=agg_num, column=9, value=long_s)
                count_sheet.cell(row=agg_num, column=10, value=count)

                count_sheet.cell(row=agg_num, column=13, value=longitude)
                count_sheet.cell(row=agg_num, column=14, value=latitude)
                agg_num = agg_num + 1
        workbook.save(self.data_path)

        # for flow_index in range(index, sheet_max_row + 1):
        #     if judge == 0 and longitude == str(sheet.cell(row=flow_index, column=8).value) and latitude == str(
        #             sheet.cell(row=flow_index, column=9).value):
        #         title = title + "@@" + str(sheet.cell(row=flow_index, column=2).value)
        #         date = date + "@@" + str(sheet.cell(row=flow_index, column=3).value)
        #         mode = mode + "@@" + str(sheet.cell(row=flow_index, column=6).value)
        #         short_s = short_s + "@@" + str(sheet.cell(row=flow_index, column=4).value)
        #         long_s = long_s + "@@" + str(sheet.cell(row=flow_index, column=5).value)
        #         site = site + "@@" + str(sheet.cell(row=flow_index, column=7).value)

        # maxrow = count_sheet.max_row
        # if count_sheet.max_row == 0:
        #     count_sheet.cell(row=1, column=1, value=id)
        #     count_sheet.cell(row=1, column=2, value=site)
        #     count_sheet.cell(row=1, column=3, value=longitude)
        #     count_sheet.cell(row=1, column=4, value=latitude)
        #     count_sheet.cell(row=1, column=5, value=mode)
        #     count_sheet.cell(row=1, column=6, value=title)
        #     count_sheet.cell(row=1, column=7, value=date)
        #     count_sheet.cell(row=1, column=8, value=short_s)
        #     count_sheet.cell(row=1, column=9, value=long_s)
        #     count_sheet.cell(row=1, column=10, value=1)
        #     workbook.save(self.data_path)
        # else:
        #     for c_index in range(1, maxrow + 1):
        #         if longitude == str(count_sheet.cell(row=c_index, column=3).value) and latitude == str(
        #                 count_sheet.cell(row=c_index, column=4).value):
        #             # mode
        #             count_sheet.cell(row=1, column=5,
        #                              value=str(count_sheet.cell(row=c_index, column=5).value) + "@@" + mode)
        #             # title
        #             count_sheet.cell(row=1, column=6,
        #                              value=str(count_sheet.cell(row=c_index, column=6).value) + "@@" + title)
        #             # date
        #             count_sheet.cell(row=1, column=7,
        #                              value=str(count_sheet.cell(row=c_index, column=7).value) + "@@" + date)
        #             # short-s
        #             count_sheet.cell(row=1, column=8,
        #                              value=str(count_sheet.cell(row=c_index, column=8).value) + "@@" + short_s)
        #             # long-s
        #             count_sheet.cell(row=1, column=9,
        #                              value=str(count_sheet.cell(row=c_index, column=9).value) + "@@" + long_s)
        #             count_sheet.cell(row=1, column=10,
        #                              value=int(str(count_sheet.cell(row=c_index, column=10).value)) + 1)
        #
        #         else:
        #             count_sheet.cell(row=maxrow + 1, column=1, value=id)
        #             count_sheet.cell(row=maxrow + 1, column=2, value=site)
        #             count_sheet.cell(row=maxrow + 1, column=3, value=longitude)
        #             count_sheet.cell(row=maxrow + 1, column=4, value=latitude)
        #             count_sheet.cell(row=maxrow + 1, column=5, value=mode)
        #             count_sheet.cell(row=maxrow + 1, column=6, value=title)
        #             count_sheet.cell(row=maxrow + 1, column=7, value=date)
        #             count_sheet.cell(row=maxrow + 1, column=8, value=short_s)
        #             count_sheet.cell(row=maxrow + 1, column=9, value=long_s)
        #             count_sheet.cell(row=maxrow + 1, column=10, value=1)


cmn = CountModeNums()
cmn.processes()
