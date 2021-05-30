import openpyxl
import re


class ModeExtraction:

    def __init__(self):
        # data path
        self.data_path = '../data/ECM_data_all.xlsx'

        # dict files
        self.trigger_words_path = '../data/trigger_words.txt'

        self.trigger_words = self.load_dict_file(self.trigger_words_path)

        # pattern files
        self.mode_pattern_path = '../data/mode_RE.txt'

        self.mode_patterns = self.load_dict_file(self.mode_pattern_path)

        # split_mode files
        self.split_mode_path = '../data/split_modes.txt'

        self.split_mode_patterns = self.load_dict_file(self.split_mode_path)

        # existing_modes files
        self.existing_mode_path = '../data/dict_argri_modes.txt'

        self.current_modes = self.load_dict_file(self.existing_mode_path)

        # filter words files
        self.filter_dict_path = '../data/dict_filter.txt'

        self.filter_words = self.load_dict_file(self.filter_dict_path)

        # all argri_modes
        self.argri_modes = set()
        self.argri_modes_list = []

    def extract_file(self):
        # step 1: 载入数据
        workbook = self.load_excel(self.data_path)
        sheet_list = self.load_sheet_list(workbook)

        # step 2: 抽取模式/句子
        for sheet in sheet_list:
            content_list = self.get_contentlist_from_sheet(sheet)
            row = 2
            for content in content_list:
                # potential sentences
                modes = self.extract_potential_from_content(content)
                sheet.cell(row=row, column=9, value=str(modes))

                # modes
                modes_with_sentences, modes = self.extract_mode_from_content(content, sheet, row)
                sheet.cell(row=row, column=10, value=str(modes_with_sentences))
                sheet.cell(row=row, column=11, value=str(modes))
                row = row + 1

            print(sheet.title + ' has been processed.')
            workbook.save(self.data_path)

        # step 3： 存储模式词表sheet
        self.creat_modes_list_sheet(workbook)

    def extract_mode_from_content(self, content, sheet, row):
        modes_withsentence = []
        modes_set = set()
        modes = []
        print('content:' + content[0:3] + '   sheet name: ' + sheet.title + ' row: ' + str(row))
        sentences = self.cut_content_into_sentences(content)
        trigger_sentences = []
        # 获取包含触发词句子
        for sentence in sentences:
            if self.contain_trigger_word(sentence, self.trigger_words):
                part_sentences = sentence.split('，')
                for ps in part_sentences:
                    trigger_sentences.append(ps)
        # 从触发词句子，匹配模式
        for ts in trigger_sentences:
            existing_mode_cursor = 0
            for pattern in self.mode_patterns:
                temp_mode = re.compile(pattern).findall(ts)
                # 匹配到模式
                if temp_mode:
                    m = str(temp_mode[0][0])
                    multi_modes = self.split_existing_modes(m)
                    if multi_modes:     # 多重模式需要分割
                        for each_m in multi_modes:
                            if self.is_filter_word(each_m) is False:
                                modes_withsentence.append(each_m + '@@' + ts)
                                self.argri_modes.add(each_m)
                                self.argri_modes_list.append(each_m)
                                modes_set.add(each_m)
                                existing_mode_cursor = 1
                    else:
                        if self.is_filter_word(modes_withsentence) is False:
                            modes_withsentence.append(m + '@@' + ts)
                            self.argri_modes.add(m)
                            self.argri_modes_list.append(m)
                            modes_set.add(m)
                            existing_mode_cursor = 1
            # 未匹配到模式，采用词表匹配，兜底
            if existing_mode_cursor == 0:
                for existing_mode in self.current_modes:
                    if existing_mode in ts:
                        modes_withsentence.append(existing_mode + '@@' + ts)
                        self.argri_modes.add(existing_mode)
                        self.argri_modes_list.append(existing_mode)
                        modes_set.add(existing_mode)
        for m_set in modes_set:
            modes.append(m_set)

        return modes_withsentence, modes

    def is_filter_word(self, mode):
        belong_filter_list = False
        for fw in self.filter_words:
            if fw in mode:
                belong_filter_list = True
        return belong_filter_list

    def split_existing_modes(self, mode_str):
        split_modes = []
        for smode in self.split_mode_patterns:
            if smode in mode_str:
                split_modes = mode_str.split(smode)
        return split_modes

    def extract_potential_from_content(self, content):
        potential_sents = []
        sentences = self.cut_content_into_sentences(content)
        trigger_sentences = []
        for sentence in sentences:
            if self.contain_trigger_word(sentence, self.trigger_words):
                trigger_sentences.append(sentence)
        for ts in trigger_sentences:
            potential_sents.append(ts)
        return potential_sents

    def contain_trigger_word(self, sentence, trigger_word_list):
        contain = False
        for tw in trigger_word_list:
            if tw in sentence:
                contain = True
        return contain

    def cut_content_into_sentences(self, content):
        sentences = []
        para = re.sub('([。！？\?])([^”’])', r"\1\n\2", content)  # 单字符断句符
        para = re.sub('(\.{6})([^”’])', r"\1\n\2", para)  # 英文省略号
        para = re.sub('(\…{2})([^”’])', r"\1\n\2", para)  # 中文省略号
        para = re.sub('([。！？\?][”’])([^，。！？\?])', r'\1\n\2', para)
        # 如果双引号前有终止符，那么双引号才是句子的终点，把分句符\n放到双引号后，注意前面的几句都小心保留了双引号
        para = para.strip()  # 段如果有多余的\n就去掉它
        # 很多规则中会考虑分号;，但是这里我把它忽略不计，破折号、英文双引号等同样忽略，需要的再做些简单调整即可。

        # 再精细化处理
        temp_sentences = para.split("\n")
        for s in temp_sentences:
            # 过滤空字符及超短句
            if s != '' and len(s) > 5:
                sentences.append(s.strip())
        return sentences

    def load_excel(self, file_path):
        try:
            wb = openpyxl.load_workbook(file_path)
            print("data file read successfully.")
            return wb
        except Exception as res:
            print("data file read unsuccessfully.")
            print(res)

    def load_sheet_list(self, wb):
        sheets = []
        if wb:
            shs = wb.get_sheet_names()
            for i in range(len(shs)):
                sheets.append(wb.get_sheet_by_name(shs[i]))
        return sheets

    def get_contentlist_from_sheet(self, sheet):
        content_list = []
        for i in range(sheet.max_row + 1):
            if i > 1:
                # read each row
                content_list.append(str(sheet.cell(row=i, column=7).value))
        return content_list

    def load_dict_file(self, path):
        mode_re_list = []
        try:
            file = open(path, 'r', encoding='utf8')

            for pattern in file.readlines():
                mode_re_list.append(pattern.strip())
            mode_re_list.pop(0)
        except:
            print("load dict file error!")
        return mode_re_list

    def creat_modes_list_sheet(self, workbook):
        modes_list_sheet = workbook.create_sheet('modes_list', 0)
        word_num = 1
        for mode_each_word in self.argri_modes:
            # name
            modes_list_sheet.cell(row=word_num, column=1, value=str(mode_each_word))
            # 次数
            mode_occur_times = self.argri_modes_list.count(mode_each_word)
            modes_list_sheet.cell(row=word_num, column=2, value=str(mode_occur_times))
            word_num = word_num + 1
        workbook.save(self.data_path)


me = ModeExtraction()
me.extract_file()
