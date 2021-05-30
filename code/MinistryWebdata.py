from goose3 import Goose
from goose3.text import StopWordsChinese
import requests
from fake_useragent import UserAgent
from lxml import etree
import openpyxl
import re


class MinistryWebdata:
    def __init__(self):
        self.ua = UserAgent()

        # 农业模式
        self.ministryweburl = 'http://www.moa.gov.cn/was5/web/search?channelid=233424&searchword=%E5%86%9C%E4%B8%9A%E6%A8%A1%E5%BC%8F&perpage=&orderby=-DOCRELTIME&searchscope=&timestart=&timeend=&chnls=&andsen=&total=&orsen=%E5%86%9C%E4%B8%9A%E6%A8%A1%E5%BC%8F&exclude=&page='


        # xlsx
        self.wb = openpyxl.load_workbook('../data/MinistryText.xlsx')
        self.MinistryTexts_Sheet1 = self.wb['Sheet1']
        self.max_row = self.MinistryTexts_Sheet1.max_row

        self.article_link_list = []
        self.article_date_list = []

        self.g = Goose({'browser_user_agent': 'Mozilla', 'stopwords_class': StopWordsChinese})

    # step 1
    def crawl_ministry_web(self):
        for i in range(1, 56):
            self.get_text_link(self.ministryweburl + str(i))
            print("finish page " + str(i) + " link crawling.")

        self.write_article_link_to_xlsx()

    # step 2
    def crawl_each_article(self):
        # read links from the xlsx
        for i in range(2, self.max_row):
            each_link = self.MinistryTexts_Sheet1.cell(row=i, column=2).value
            self.article_link_list.append(each_link)

        # crawl information from each link
        linknum = 2
        for link in self.article_link_list:
            print("extracting title and content from the link of " + str(linknum - 1))
            eachtitle, eachcontent = self.crawl_information_from_each_link(link)
            self.MinistryTexts_Sheet1.cell(row=linknum, column=3, value=eachtitle)
            self.MinistryTexts_Sheet1.cell(row=linknum, column=7, value=eachcontent)
            linknum = linknum + 1
        self.wb.save("../data/MinistryText.xlsx")
        print("title and content have been extracted successfully!")

    # step 3
    def crawl_title_and_date(self):
        for i in range(1, 56):
            self.get_text_date(self.ministryweburl + str(i))

        i = 1
        for date in self.article_date_list:
            i = i + 1
            print("finish " + str(i) + " date crawling.")
            self.MinistryTexts_Sheet1.cell(row=i, column=4, value=date)
        self.wb.save("../data/MinistryText.xlsx")


    def crawl_information_from_each_link(self, url):
        # information to be extracted
        extract_title = ''
        extract_date = ''
        extract_authors = ''
        extract_datasource = ''
        extract_content = ''

        headers = {
            "User-Agent": self.ua.random,
            'Connection': 'close',
        }
        content = ''
        try:
            res = requests.get(url, headers=headers, timeout=8)
        except:
            content = url
            print('未获取链接！')
        if content != url:
            print(res.url)

            tree = etree.HTML(res.content)

            # title
            title_list = tree.xpath(
                '//h1[@class="bjjMTitle"]|//h1[@class="sj_xiang_biaoti"]|//h1[@class="ft_24 ft_w_n txt_c mg_t_35 mg_b_35"]|//p[@class="detail_title"]')
            if title_list != []:
                for t in title_list:
                    title = t.xpath("text()")
                    tt = str(title)[2:len(title) - 3]
                    # print("title:"+tt)
                extract_title = tt

            # content _goose3解析
            article = self.g.extract(url)
            extract_content = article.cleaned_text
            # print("content: "+content)

        return extract_title, extract_content

    def get_text_link(self, url):
        headers = {
            "User-Agent": self.ua.random,
            'Connection': 'close',
        }
        content = ''
        try:
            res = requests.get(url, headers=headers, timeout=8)
        except:
            content = url
            print('未获取链接！')
        if content != url:
            print(res.url)
            tree = etree.HTML(res.text)
            article_list_inwebpage = tree.xpath('//div[@class="tit"]//a[@class="dcotitlename"]')
            for a in article_list_inwebpage:
                each_article_link = a.xpath("@href")
                self.article_link_list.append(each_article_link)

    def get_text_date(self, url):
        headers = {
            "User-Agent": self.ua.random,
            'Connection': 'close',
        }
        content = ''
        try:
            res = requests.get(url, headers=headers, timeout=8)
        except:
            content = url
            print('未获取链接！')
        if content != url:
            print(res.url)
            tree = etree.HTML(res.content)
            date_list = tree.xpath('//span[@class="fbsj"]')
            for d in date_list:
                date_text = d.xpath("text()")
                self.article_date_list.append(str(date_text)[7:len(str(date_text))-25])

    def write_article_link_to_xlsx(self):
        i = 1
        for link in self.article_link_list:
            i = i + 1

            self.MinistryTexts_Sheet1.cell(row=i, column=2, value=str(link)[2:len(str(link)) - 2])
        self.wb.save("../data/MinistryText.xlsx")
