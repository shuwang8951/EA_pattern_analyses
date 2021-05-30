from goose3 import Goose
from goose3.text import StopWordsChinese
import requests
from fake_useragent import UserAgent
from lxml import etree
import openpyxl
import re
from EcoCivMdl.code import News


class YangshiNewsdata:
    def __init__(self):
        self.ua = UserAgent()

        # goose网页正文获取
        self.g = Goose({'browser_user_agent': 'Mozilla', 'stopwords_class': StopWordsChinese})

        # 农业模式
        self.query_url_pre = 'https://search.cctv.com/search.php?qtext='
        self.query_url_suffix = '&sort=relevance&type=web&vtime=&datepid=1&channel=&page='

        # xlsx
        self.wb = openpyxl.load_workbook('../data/YangshiNewsText.xlsx')
        self.YangshiNewsTexts_Sheet1 = self.wb['Sheet1']

        # query word list
        self.qwwb = openpyxl.load_workbook('../data/生态农业概念分类与词汇表.xlsx')
        self.zhongzhiye_querywords_sheet = self.qwwb['生态种植业词表']
        self.zhongzhiye_query_word_list = self.zhongzhiye_querywords_sheet['A']
        # for w in self.zhongzhiye_query_word_list:
        #     print(w.value)

    # logic
    def crawl_yangshinews_web(self):
        index = 1
        for each_query_word_cell in self.zhongzhiye_query_word_list:
            print("-----------start crawl word:" + each_query_word_cell.value + " news.-----------" + str(
                index) + "/" + str(len(self.zhongzhiye_query_word_list)))
            self.get_news_from_query_word(each_query_word_cell.value)
            print("-----------finish crawl word:" + each_query_word_cell.value + " news.-----------" + str(
                index) + "/" + str(len(self.zhongzhiye_query_word_list)))
            index = index + 1

    # calculate max page for each query word
    def max_page_calculator_by_query_word(self, query_word):
        # initial setting
        max_page = 1

        query_url = self.query_url_pre + query_word + " 生态 农业" + self.query_url_suffix + '1'
        tree = self.get_HTML_tree_from_url(query_url)
        # parse the max number of the search results
        max_number_section = tree.xpath('//div[@class="lmdhd"]//span')
        items = []
        if max_number_section:
            max_number_description = max_number_section[0].xpath("text()")
            search_item_pattern = '''结果共(.*?)条'''
            items = re.compile(search_item_pattern).findall(max_number_description[0])
        if str.isdigit(items[0]) == True:
            result_num = int(items[0])
        else:
            result_num = 0
        if result_num > 300:
            max_page = 30
        else:
            max_page = result_num // 10
            if max_page < 1:
                max_page = 1
        print(query_word + " max page: " + str(max_page))
        return max_page

    # get the html tree structure from the url link
    def get_HTML_tree_from_url(self, url):
        headers = {
            "User-Agent": self.ua.random,
            'Connection': 'close',
        }
        content = ''
        tree = etree.HTML('')
        try:
            res = requests.get(url, headers=headers, timeout=8)
        except:
            content = url
            print('未获取链接！')
        if content != url:
            # print(res.url)
            tree = etree.HTML(res.text)
        return tree

    def get_news_from_query_word(self, query_word):

        max_page = self.max_page_calculator_by_query_word(query_word)
        for i in range(1, max_page):
            # news in each page
            news_list = []
            url = self.query_url_pre + query_word + self.query_url_suffix + str(i)
            print(url)
            tree = self.get_HTML_tree_from_url(url)

            url = []
            title = []
            content = []
            date = []

            # url
            url_section = tree.xpath('//li//h3[@class="tit"]//span')
            for u_section in url_section:
                each_news_link = u_section.xpath("@lanmu1")
                if each_news_link:
                    url.append(each_news_link[0])
                else:
                    url.append('')

            # title
            title_section = tree.xpath('//h3//span//a')
            for t_section in title_section:
                each_news_title = t_section.xpath('text()')
                title.append("".join(each_news_title))

            # date
            date_section = tree.xpath('//div[@class="src-tim"]//span[@class="tim"]')
            for d_section in date_section:
                each_news_date = d_section.xpath('text()')
                nd = each_news_date[0][5:len(each_news_date) - 9]
                date.append(nd)

            # content
            for link in url:
                news_content = "null"
                try:
                    # content _goose3解析
                    news_article = self.g.extract(link)
                    news_content = news_article.cleaned_text
                except Exception as result:
                    print(result)
                finally:
                    content.append(news_content)

            for index in range(0, len(url)):
                if url[index] and title[index] and date[index] and content[index]:
                    news = News.News(url[index], title[index], date[index], content[index], query_word)
                    news_list.append(news)
            print("get -" + query_word + "- page" + str(i) + " news of " + str(len(news_list)))

            self.write_news_to_xlsx(news_list)

    def write_news_to_xlsx(self, news_list):

        existing_row_num = self.YangshiNewsTexts_Sheet1.max_row

        insert_news_num = len(news_list)
        for index in range(0, insert_news_num):
            news = news_list[index]

            # url
            self.YangshiNewsTexts_Sheet1.cell(row=existing_row_num + 1 + index, column=2, value=news.url)
            # title
            self.YangshiNewsTexts_Sheet1.cell(row=existing_row_num + 1 + index, column=3, value=news.title)
            # date
            self.YangshiNewsTexts_Sheet1.cell(row=existing_row_num + 1 + index, column=4, value=news.date)
            # content
            self.YangshiNewsTexts_Sheet1.cell(row=existing_row_num + 1 + index, column=7, value=news.content)
            # query_word
            self.YangshiNewsTexts_Sheet1.cell(row=existing_row_num + 1 + index, column=6, value=news.query_word)

        self.wb.save("../data/YangshiNewsText.xlsx")
