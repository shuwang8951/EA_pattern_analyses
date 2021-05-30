import openpyxl
import pynlpir
import re


class MinistryDataParse:
    def __init__(self):
        # xlsx
        self.wb = openpyxl.load_workbook('../data/MinistryText.xlsx')
        self.MinistryTexts_Sheet1 = self.wb['Sheet1']

        self.content_list = self.read_col(self.MinistryTexts_Sheet1['G'])

        pynlpir.open()

        # load mode_RE_list
        self.mode_re_list = self.load_mode_RE('../data/mode_RE.txt')

        self.parse()

    def read_col(self, row_name):
        datalist = []
        for x in range(len(row_name)):
            if x > 0:
                value = row_name[x].value
                if value is None:
                    pass
                else:
                    value = str(value)
                    datalist.append(value)
        return datalist

    def parse(self):
        # self.extract_location()
        self.extract_mode()

    def extract_location(self):
        index = 2
        for content in self.content_list:
            toponyms = []
            max_toponym = ""
            max_count = 0
            max2_toponym = ""
            max2_count = 0

            seg_list = pynlpir.segment(content, pos_names='raw')
            for word_pair in seg_list:
                if word_pair[1] == "ns":
                    toponyms.append(word_pair[0].replace("\n", ""))

            if toponyms != []:

                for toponym in toponyms:
                    c = toponyms.count(toponym)
                    if c > max_count:
                        max_count = c
                        max_toponym = toponym

                    elif c > max2_count and toponym != max_toponym:
                        max2_count = c
                        max2_toponym = toponym
            else:
                toponyms.append("")
            print("--------------------------")
            print(toponyms)
            print(max_toponym)
            print(max2_toponym)
            self.MinistryTexts_Sheet1.cell(row=index, column=8, value=str(toponyms))
            self.MinistryTexts_Sheet1.cell(row=index, column=9, value=max_toponym + "," + max2_toponym)
            index = index + 1
        self.wb.save("../data/MinistryText.xlsx")

    def load_mode_RE(self, path):
        mode_re_list = []
        try:
            file = open(path, 'r', encoding='utf8')

            for pattern in file.readlines():
                mode_re_list.append(pattern.strip())
        except:
            print("load mode_RE error!")
        return mode_re_list

    def extract_mode(self):
        index = 1

        # mode = ""
        # sentences = re.split('''。|！|？''', self.content_list[235])
        # i = 0
        # for sen in sentences:
        #     print(str(i))
        #     i = i + 1
        #     print(sen + "。")

        for content in self.content_list:
            mode = ''
            modes_list=[]
            sentences = re.split('''。|！|？''', content)
            modes = set()
            for sen in sentences:
                for mode_pattern in self.mode_re_list:
                    temp = re.compile(mode_pattern).findall(sen)
                    # mode_combine = ''
                    if temp:
                        m = str(temp[0][0])
                        modes.add(m)
                        # mode_combine = mode_combine + m
                        # for em in temp:
                        #     modes.add(str(em))
                        #     mode_combine = mode_combine + str(em)
                        # print(temp[0][0])
            for each_mode in modes:
                if each_mode != '' or each_mode != '“':
                    mode = mode + str(each_mode) + '；'
                    modes_list.append(str(each_mode)[1:len(str(each_mode))-1])
            print("extract modes from " + str(index) + " text...")
            index = index + 1
            self.MinistryTexts_Sheet1.cell(row=index, column=10, value=str(modes_list))
            # if index == 3:
            #     break
        self.wb.save("../data/MinistryText.xlsx")
